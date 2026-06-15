import json
import argparse
import os
import sys
from datetime import datetime, date, timedelta
from collections import OrderedDict

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA_DIR = os.path.join(ROOT, "data")


def load_json(filename):
    path = os.path.join(DATA_DIR, filename)
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        r = data.get("result", data)
        if isinstance(r, dict):
            return r.get("result", r)
        return r
    return data


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.replace("Z", "")
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


def parse_date_only(val):
    d = parse_date(val)
    if d:
        return d.date()
    return None


MONTH_NAMES_RU = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель",
    5: "май", 6: "июнь", 7: "июль", 8: "август",
    9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь",
}

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
THICK = Side(style="thick", color="000000")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GANTT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
GANTT_ACTUAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_STATUS_IDS = {3, 5, 6}  # Выполнено, Ожидание запуска снабжения, Снабжение
FONT_WHITE = Font(name="Calibri", size=7)
HEADER_FONT = Font(name="Calibri", size=11, bold=False)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
DATA_FONT = Font(name="Calibri", size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_gantt_excel(vitrage_data, sprav_data, zadanie_data, object_id, corpus_filter, section_filter, start_date, end_date, output_path):
    items = vitrage_data

    if isinstance(items, dict) and "result" in items:
        items = items["result"]
    if isinstance(items, dict) and "result" in items:
        items = items["result"]

    items = [r for r in items if not r.get("__deletedAt")]

    if object_id:
        items = [r for r in items if r.get("itogovyi_id") == object_id]
    if corpus_filter:
        items = [r for r in items if r.get("korpus") == corpus_filter]
    if section_filter:
        items = [r for r in items if r.get("sections") == section_filter]

    if not items:
        print("Нет данных по указанным фильтрам")
        return

    zadanie_lookup = {}
    if zadanie_data:
        if isinstance(zadanie_data, dict) and "result" in zadanie_data:
            zitems = zadanie_data["result"]
            if isinstance(zitems, dict) and "result" in zitems:
                zitems = zitems["result"]
        else:
            zitems = zadanie_data
        for z in zitems:
            zid = z.get("__id")
            if zid:
                zadanie_lookup[zid] = z

    dop_zadanie_ids = set()
    for zid, zrec in zadanie_lookup.items():
        if zrec.get("dop_zakaz_kmd_ref"):
            dop_zadanie_ids.add(zid)
    if dop_zadanie_ids:
        print(f"Исключено доп-заданий КМД (dop_zakaz_kmd_ref): {len(dop_zadanie_ids)}")

    status_data = load_json("statusy_zadanie_na_kmd.json")
    status_items = {}
    if status_data and "statusItems" in status_data:
        for s in status_data["statusItems"]:
            status_items[s["id"]] = s

    users_data = load_json("users.json")
    users_lookup = {}
    if isinstance(users_data, dict) and "result" in users_data:
        uitems = users_data["result"]
        if isinstance(uitems, dict) and "result" in uitems:
            uitems = uitems["result"]
    else:
        uitems = users_data
    for u in uitems:
        uid = u.get("__id")
        if uid:
            users_lookup[uid] = u

    object_name = object_id or ""
    for r in items:
        if r.get("itogovyi_id"):
            object_name = r.get("itogovyi_id", object_name)
            break

    sprav_name = object_name
    for r in sprav_data:
        if r.get("itogovyi_id") == object_name:
            sprav_name = r.get("kodovoe_nazvanie") or r.get("__name", object_name)
            break

    records = []
    for r in items:
        vitrazh_str = r.get("vitrazh", "")
        if not vitrazh_str:
            continue
        names = [v.strip() for v in vitrazh_str.split(",") if v.strip()]
        if not names:
            continue
        nachalo = parse_date(r.get("nachalo_kmd"))
        konec = parse_date(r.get("konec_kmd"))
        fakt_nachalo = parse_date(r.get("fakt_nachalo_kmd")) or nachalo
        fakt_konec = parse_date(r.get("fakt_konec_kmd")) or konec

        korpus = r.get("korpus", "")
        section = r.get("sections", "")
        ploshad = r.get("ploshad_po_km_m2")
        stvorok = r.get("kolichestvo_stvorok")
        ventresh = r.get("kolichestvo_ventreshetok")

        main_zad_ids = set(r.get("zadanie_na_kmd_2025") or [])
        dop_zad_ids = set(r.get("dop_kmd") or [])
        only_main_zad_ids = main_zad_ids - dop_zadanie_ids

        is_freelance = False
        all_zad_ids = only_main_zad_ids | dop_zad_ids
        for zid in all_zad_ids:
            zrec = zadanie_lookup.get(zid)
            if zrec and zrec.get("freelance"):
                is_freelance = True
                break

        constructor_initials = ""
        if is_freelance:
            constructor_initials = "Fr"
        else:
            cons_ids = r.get("constructor_kmd") or []
            if isinstance(cons_ids, str):
                cons_ids = [cons_ids]
            for cid in cons_ids:
                u = users_lookup.get(cid)
                if u:
                    fn = u.get("fullname")
                    if isinstance(fn, dict):
                        a = (fn.get("lastname") or "")[:1].upper()
                        b = (fn.get("firstname") or "")[:1].upper()
                        constructor_initials = a + b
                    else:
                        name = fn or u.get("__name", "")
                        if isinstance(name, str):
                            parts = name.split()
                            if len(parts) >= 2:
                                constructor_initials = parts[0][0].upper() + parts[1][0].upper()
                            elif parts:
                                constructor_initials = parts[0][0].upper()
                    break

        rop_end = None
        for zid in only_main_zad_ids:
            zrec = zadanie_lookup.get(zid)
            if not zrec:
                continue
            ro = parse_date(zrec.get("rop_date_original"))
            if ro and not rop_end:
                rop_end = ro.date()

        raw_fakt_konec = parse_date(r.get("fakt_konec_kmd"))

        zadanie_status_id = None
        for zid in only_main_zad_ids:
            zrec = zadanie_lookup.get(zid)
            if zrec:
                st = zrec.get("__status")
                if isinstance(st, dict):
                    zadanie_status_id = st.get("status")
                    break

        n = len(names)
        for i, name in enumerate(names):
            records.append({
                "name": name,
                "korpus": korpus,
                "section": section,
                "nachalo": nachalo.date() if nachalo else None,
                "fakt_konec": fakt_konec.date() if fakt_konec else None,
                "has_fakt_konec": raw_fakt_konec is not None,
                "rop_end": rop_end,
                "zadanie_status_id": zadanie_status_id,
                "initials": constructor_initials,
                "ploshad": ploshad / n if ploshad and n > 0 else ploshad,
                "stvorok": stvorok / n if stvorok and n > 0 else stvorok,
                "ventresh": ventresh / n if ventresh and n > 0 else ventresh,
            })

    if not records:
        print("Нет витражей для отображения")
        return

    all_dates = []
    for rec in records:
        for d in [rec["nachalo"], rec["fakt_konec"], rec["rop_end"]]:
            if d:
                all_dates.append(d)

    global_min = min(all_dates) if all_dates else date.today()
    global_max = max(all_dates) if all_dates else date.today()

    if start_date:
        try:
            global_min = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            pass
    if end_date:
        try:
            global_max = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            pass

    months = OrderedDict()
    d = global_min
    while d <= global_max:
        key = (d.year, d.month)
        if key not in months:
            months[key] = {"start": d, "end": d}
        else:
            months[key]["start"] = min(months[key]["start"], d)
            months[key]["end"] = max(months[key]["end"], d)
        d += timedelta(days=1)

    for key in months:
        m = months[key]
        days_in_month = (date(m["start"].year, m["start"].month % 12 + 1, 1) - timedelta(days=1)).day if m["start"].month < 12 else 31
        m["days"] = list(range(1, days_in_month + 1))

    records.sort(key=lambda r: (r["korpus"] or "", r["section"] or "", r["name"] or ""))

    corpus_groups = []
    current_corpus = None
    for rec in records:
        cor_key = rec["korpus"]
        if cor_key != current_corpus:
            corpus_groups.append({"korpus": cor_key, "sections": []})
            current_corpus = cor_key
        sec_key = rec["section"]
        if not corpus_groups[-1]["sections"] or corpus_groups[-1]["sections"][-1]["section"] != sec_key:
            corpus_groups[-1]["sections"].append({"section": sec_key, "records": []})
        corpus_groups[-1]["sections"][-1]["records"].append(rec)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График производства"
    ws.freeze_panes = "E6"

    col = 1
    INFO_COLS = {
        "A": 1, "B": 2, "C": 3, "D": 4, "E": 5,
        "F": 6, "G": 7, "H": 8, "I": 9, "J": 10,
    }
    INFO_HEADERS = [
        ("A", "корпус"),
        ("B", "Секция"),
        ("C", "№ витража по КМ"),
        ("D", "№ витража по АР"),
        ("E", "наименование работ"),
        ("F", "Этажность."),
        ("G", "Створки по КМ, шт."),
        ("H", "Вентреш. по КМ, шт."),
        ("I", "КМ, в м\u00b2"),
        ("J", "кол-во, шт"),
    ]

    day_cols = []
    for mkey, minfo in months.items():
        year, month_num = mkey
        for day_num in minfo["days"]:
            day_cols.append({
                "date": date(year, month_num, day_num),
                "month_key": mkey,
            })

    gantt_end_col_num = len(INFO_COLS) + len(day_cols)
    MAX_COL = INFO_COLS["J"] + len(day_cols)

    # Address merge removed
    # Address cell removed
    # Formatting removed

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=MAX_COL)
    proj_cell = ws.cell(row=2, column=1, value=f"{sprav_name} ({object_name})")
    proj_cell.font = Font(name="Calibri", size=12, bold=True)
    proj_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 28

    header_style_border = Border(
        top=MEDIUM, bottom=THIN, left=THIN, right=THIN
    )
    thin_border = Border(
        top=None, bottom=THIN, left=THIN, right=THIN
    )
    thin_border_bottom = Border(
        top=THIN, bottom=THIN, left=THIN, right=THIN
    )

    for col_letter, hdr_text in INFO_HEADERS:
        ci = INFO_COLS[col_letter]
        ws.merge_cells(start_row=4, start_column=ci, end_row=5, end_column=ci)
        cell = ws.cell(row=4, column=ci, value=hdr_text)
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        cell.border = header_style_border

    for mkey, minfo in months.items():
        year, month_num = mkey
        mnth_name = MONTH_NAMES_RU[month_num]
        start_col = INFO_COLS["J"] + 1
        for dc in day_cols:
            if dc["month_key"] == mkey:
                break
            start_col += 1
        end_col = start_col + len(minfo["days"]) - 1
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        cell = ws.cell(row=3, column=start_col, value=f"{mnth_name} {year}")
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = CENTER

    day_cell_border = Border(
        top=THIN, bottom=THIN, left=THIN, right=THIN
    )

    def border_for_date(d):
        left = THIN
        if d.day == 1:
            left = THICK
        elif d.weekday() == 0:
            left = MEDIUM
        return Border(top=THIN, bottom=THIN, left=left, right=THIN)

    for i, dc in enumerate(day_cols):
        col_idx = INFO_COLS["J"] + 1 + i
        day_date = dc["date"]
        day_num = day_date.day
        cell_border = border_for_date(day_date)
        parity = day_num % 2
        row_4 = 4 if parity == 1 else 5
        row_5 = 5 if parity == 1 else 4
        cell_top = ws.cell(row=4, column=col_idx)
        cell_bot = ws.cell(row=5, column=col_idx)
        if row_4 == 4:
            cell_top.value = day_num
            cell_bot.value = ""
        else:
            cell_top.value = ""
            cell_bot.value = day_num
        cell_top.font = Font(name="Calibri", size=8)
        cell_bot.font = Font(name="Calibri", size=8)
        cell_top.alignment = CENTER
        cell_bot.alignment = CENTER
        cell_top.border = cell_border
        cell_bot.border = cell_border

    FIRST_DATA_ROW = 6
    current_row = FIRST_DATA_ROW
    total_border = Border(
        top=THIN, bottom=THIN, left=THIN, right=THIN
    )

    for cor_idx, cor_group in enumerate(corpus_groups):
        korpus = cor_group["korpus"]
        cor_start_row = current_row

        for sec_idx, sec_group in enumerate(cor_group["sections"]):
            section = sec_group["section"]
            recs = sec_group["records"]
            start_row = current_row

            for rec_idx, rec in enumerate(recs):
                row = current_row

                ws.cell(row=row, column=INFO_COLS["C"], value=rec["name"])
                ws.cell(row=row, column=INFO_COLS["D"], value=rec["name"])
                ws.cell(row=row, column=INFO_COLS["E"], value="кмд")
                ws.cell(row=row, column=INFO_COLS["F"], value="")
                if rec["stvorok"]:
                    ws.cell(row=row, column=INFO_COLS["G"], value=round(rec["stvorok"], 1))
                if rec["ventresh"]:
                    ws.cell(row=row, column=INFO_COLS["H"], value=round(rec["ventresh"], 1))
                if rec["ploshad"]:
                    ws.cell(row=row, column=INFO_COLS["I"], value=round(rec["ploshad"], 2))
                ws.cell(row=row, column=INFO_COLS["J"], value=1)

                row_fill = None
                if rec["has_fakt_konec"] and rec["fakt_konec"]:
                    if rec["zadanie_status_id"] and rec["zadanie_status_id"] not in GREEN_STATUS_IDS:
                        row_fill = YELLOW_FILL
                    else:
                        row_fill = GANTT_ACTUAL_FILL
                elif rec["nachalo"]:
                    row_fill = GANTT_FILL
                if row_fill:
                    for ci in range(INFO_COLS["C"], INFO_COLS["J"] + 1):
                        ws.cell(row=row, column=ci).fill = row_fill

                for ci in range(1, MAX_COL + 1):
                    c = ws.cell(row=row, column=ci)
                    c.font = DATA_FONT
                    c.alignment = CENTER
                    c.border = Border(
                        top=MEDIUM if rec_idx == 0 else THIN,
                        bottom=THIN, left=THIN, right=THIN
                    )

                for di, dc in enumerate(day_cols):
                    col_idx = INFO_COLS["J"] + 1 + di
                    day_date = dc["date"]
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = border_for_date(day_date)

                    is_colored = False
                    if rec["nachalo"]:
                        start_n = rec["nachalo"]
                        if rec["has_fakt_konec"] and rec["fakt_konec"]:
                            end_n = rec["fakt_konec"]
                            if rec["zadanie_status_id"] and rec["zadanie_status_id"] not in GREEN_STATUS_IDS:
                                fill = YELLOW_FILL
                            else:
                                fill = GANTT_ACTUAL_FILL
                        elif rec["rop_end"]:
                            end_n = rec["rop_end"]
                            fill = GANTT_FILL
                        else:
                            end_n = None

                        if end_n and start_n <= day_date <= end_n:
                            cell.fill = fill
                            is_colored = True

                    if is_colored and rec["initials"]:
                        cell.value = rec["initials"]
                        cell.font = FONT_WHITE

                ws.row_dimensions[row].height = 18
                current_row += 1

            end_row = current_row - 1
            if start_row <= end_row:
                ws.merge_cells(start_row=start_row, start_column=INFO_COLS["B"], end_row=end_row, end_column=INFO_COLS["B"])
                b_cell = ws.cell(row=start_row, column=INFO_COLS["B"], value=section)
                b_cell.font = DATA_FONT
                b_cell.alignment = CENTER

            total_row = current_row
            ws.merge_cells(start_row=total_row, start_column=INFO_COLS["B"], end_row=total_row, end_column=INFO_COLS["E"])
            total_label = ws.cell(row=total_row, column=INFO_COLS["B"], value="Всего (витражи):")
            total_label.font = BOLD_FONT
            total_label.alignment = Alignment(horizontal="right", vertical="center")

            sum_stvorok = sum(r["stvorok"] or 0 for r in recs)
            sum_ventresh = sum(r["ventresh"] or 0 for r in recs)
            sum_ploshad = sum(r["ploshad"] or 0 for r in recs)
            sum_qty = len(recs)

            if sum_stvorok:
                ws.cell(row=total_row, column=INFO_COLS["G"], value=round(sum_stvorok, 1))
            if sum_ventresh:
                ws.cell(row=total_row, column=INFO_COLS["H"], value=round(sum_ventresh, 1))
            if sum_ploshad:
                ws.cell(row=total_row, column=INFO_COLS["I"], value=round(sum_ploshad, 2))
            ws.cell(row=total_row, column=INFO_COLS["J"], value=sum_qty)

            for ci in range(1, MAX_COL + 1):
                c = ws.cell(row=total_row, column=ci)
                c.font = BOLD_FONT
                c.alignment = CENTER
                c.border = total_border

            ws.row_dimensions[total_row].height = 18
            current_row += 1

        cor_end_row = current_row - 1
        if cor_start_row <= cor_end_row:
            ws.merge_cells(start_row=cor_start_row, start_column=INFO_COLS["A"], end_row=cor_end_row, end_column=INFO_COLS["A"])
            a_cell = ws.cell(row=cor_start_row, column=INFO_COLS["A"], value=korpus)
            a_cell.font = BOLD_FONT
            a_cell.alignment = CENTER

    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 19
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12

    for di in range(len(day_cols)):
        col_letter = get_column_letter(INFO_COLS["J"] + 1 + di)
        ws.column_dimensions[col_letter].width = 4

    last_data_row = current_row - 1

    ws.print_area = f"A1:{get_column_letter(MAX_COL)}{last_data_row}"
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A3

    wb.save(output_path)
    print(f"Готово: {output_path}")
    print(f"Объект: {sprav_name} ({object_name})")
    print(f"Период: {global_min} - {global_max}")
    total_sec = sum(len(g["sections"]) for g in corpus_groups)
    print(f"Корпусов: {len(corpus_groups)}, секций: {total_sec}")
    print(f"Витражей всего: {len(records)}")


def build_object_index(vitrage_data, sprav_data):
    items = vitrage_data
    if isinstance(items, dict) and "result" in items:
        items = items["result"]
    if isinstance(items, dict) and "result" in items:
        items = items["result"]

    obj_index = {}
    for r in items:
        iid = r.get("itogovyi_id")
        if not iid:
            continue
        if iid not in obj_index:
            name = iid
            for s in sprav_data:
                if s.get("itogovyi_id") == iid:
                    name = s.get("kodovoe_nazvanie") or s.get("__name", iid)
                    break
            obj_index[iid] = {"id": iid, "name": name, "korpuses": set()}
        obj_index[iid]["korpuses"].add(r.get("korpus", ""))
    return obj_index


def interactive_select(vitrage_data, sprav_data, zadanie_data):
    obj_index = build_object_index(vitrage_data, sprav_data)
    sorted_objects = sorted(obj_index.values(), key=lambda x: (x["name"], x["id"]))

    print("\nДоступные объекты:")
    for i, o in enumerate(sorted_objects, 1):
        korps = sorted(k for k in o["korpuses"] if k)
        print(f"  {i:3}. {o['name']} ({o['id']})  [{', '.join(korps)}]")

    while True:
        try:
            choice = input("\nВыберите номер: ").strip()
            if not choice:
                return
            idx = int(choice) - 1
            if 0 <= idx < len(sorted_objects):
                break
            print(f"Введите число от 1 до {len(sorted_objects)}")
        except ValueError:
            print("Введите число")

    selected = sorted_objects[idx]
    object_id = selected["id"]

    corpus_input = input("Корпус (Enter — все): ").strip()
    corpus = corpus_input if corpus_input else None

    section_input = input("Секция (Enter — все): ").strip()
    section = section_input if section_input else None

    start_input = input("Начало периода ГГГГ-ММ-ДД (Enter — авто): ").strip()
    end_input = input("Конец периода ГГГГ-ММ-ДД (Enter — авто): ").strip()
    start = start_input if start_input else None
    end = end_input if end_input else None

    output = os.path.join(ROOT, "reports", f"grafik_vitrazhey_{object_id}.xlsx")
    os.makedirs(os.path.dirname(output), exist_ok=True)

    build_gantt_excel(
        vitrage_data, sprav_data, zadanie_data,
        object_id=object_id,
        corpus_filter=corpus,
        section_filter=section,
        start_date=start,
        end_date=end,
        output_path=output,
    )


def main():
    parser = argparse.ArgumentParser(description="Генерация Excel-графика витражей")
    parser.add_argument("--object", "-o", help="itogovyi_id объекта")
    parser.add_argument("--corpus", "-k", help="Корпус (фильтр)")
    parser.add_argument("--section", "-s", help="Секция (фильтр)")
    parser.add_argument("--start", help="Начало периода (ГГГГ-ММ-ДД)")
    parser.add_argument("--end", help="Конец периода (ГГГГ-ММ-ДД)")
    parser.add_argument("--output", default=None, help="Путь к выходному файлу")

    args = parser.parse_args()

    vitrage_data = load_json("kartochka_vitrazha_po_km.json")
    sprav_data = load_json("spravochnik_id.json")
    zadanie_data = load_json("zadanie_na_kmd.json")

    if not args.object:
        interactive_select(vitrage_data, sprav_data, zadanie_data)
        return

    if args.output is None:
        safe_name = args.object.replace("/", "_").replace("\\", "_")
        args.output = os.path.join(ROOT, "reports", f"grafik_vitrazhey_{safe_name}.xlsx")

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    build_gantt_excel(
        vitrage_data, sprav_data, zadanie_data,
        object_id=args.object,
        corpus_filter=args.corpus,
        section_filter=args.section,
        start_date=args.start,
        end_date=args.end,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
